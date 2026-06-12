import openpyxl, json, os

BASE = r'C:\Users\ascha\OneDrive\Desktop\forage-data'
SRC = os.path.join(BASE, 'Kitchen Trailing 12 P5.26.xlsx')
JSON_PATH = os.path.join(BASE, 'dashboard_data.json')
PERIOD = '5'
COL = 2  # P5 in col B

METRIC_LABELS = {
    'Net Sales': '  Net Sales',
    'COGS': '  Total Cost of Goods Sold',
    'Labor': '  Total Payroll Expenses',
    'Occupancy': '  Total Occupancy',
    'EBITDA': 'EBITDA',
}
STORES = ['8001','8002','8003','8004','8005','8006','8007','8008','8009','8010']

def find_row(ws, label):
    for r in range(1, ws.max_row+1):
        v = ws.cell(r,1).value
        if v is None: continue
        if str(v).strip() == label.strip():
            return r
    return None

wb = openpyxl.load_workbook(SRC, data_only=True)
with open(JSON_PATH) as f:
    data = json.load(f)

for sid in STORES:
    sheet = sid + '-F'
    if sheet not in wb.sheetnames:
        print(f'SKIP {sheet}')
        continue
    ws = wb[sheet]
    key = sid + '_2026'
    if key not in data:
        data[key] = {m: {str(p): 0 for p in range(1,13)} for m in METRIC_LABELS}
    for metric, label in METRIC_LABELS.items():
        row = find_row(ws, label)
        if row is None:
            print(f'  {sid} {metric}: row not found')
            continue
        v = ws.cell(row=row, column=COL).value
        v = round(float(v), 2) if v is not None else 0
        data[key].setdefault(metric, {str(p):0 for p in range(1,13)})[PERIOD] = v
        print(f'{sid} {metric} P{PERIOD} (row {row}) = {v}')

with open(JSON_PATH, 'w') as f:
    json.dump(data, f, indent=2)
print('Saved', JSON_PATH)
