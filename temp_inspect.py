import openpyxl
wb = openpyxl.load_workbook('C:/Users/ascha/OneDrive/Desktop/forage-data/temp_budget.xlsx', data_only=True)

# Check State sheet first
sheets_to_check = ["State", "Hilldale", "Restaurant Operations"]
for sheet_name in sheets_to_check:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ===")
        for row in range(1, min(60, ws.max_row + 1)):
            val = ws.cell(row, 1).value
            if val:
                val_str = str(val).strip()
                # Show all rows to find the wages/salary rows
                print(f"  Row {row:3d}: {val_str}")
