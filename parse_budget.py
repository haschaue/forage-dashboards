"""Parse 2026 Budget Excel into JSON for the dashboard.

Now extracts BOTH 'Salaries/Wages Crew' and 'Total Payroll Expenses' so the
dashboard can compare R365 daily labor (crew wages) against the crew wages
budget line for an apples-to-apples comparison.
"""
import openpyxl
import json

wb = openpyxl.load_workbook('C:/Users/ascha/OneDrive/Desktop/forage-data/temp_budget.xlsx', data_only=True)

SHEET_MAP = {
    "State": "8001",
    "Hilldale": "8002",
    "Monona": "8003",
    "Middleton": "8004",
    "Champaign": "8005",
    "WFB": "8006",
    "Sun Prairie": "8007",
    "Pewaukee": "8008",
    "Public Market": "8009",
    "Brookfield": "8010",
}

PERIOD_COLS = {i+1: 3 + i*2 for i in range(12)}

def safe_round(val, digits=2):
    if val is None or isinstance(val, str):
        return 0
    try:
        return round(float(val), digits)
    except:
        return 0

budget = {}

for sheet_name, store_num in SHEET_MAP.items():
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet '{sheet_name}' not found")
        continue

    ws = wb[sheet_name]

    # Find key rows
    sales_row = None
    cogs_row = None
    payroll_row = None
    crew_wages_row = None

    for row in range(1, min(100, ws.max_row + 1)):
        val = ws.cell(row, 1).value
        if val:
            val_str = str(val).strip()
            if val_str == "2026 Sales":
                sales_row = row
            elif "Total Cost of Goods Sold" in val_str:
                cogs_row = row
            elif "Total Payroll Expenses" in val_str:
                payroll_row = row
            elif "salaries" in val_str.lower() and "crew" in val_str.lower():
                crew_wages_row = row

    budget[store_num] = {"name": sheet_name}

    for period, col in PERIOD_COLS.items():
        sales_val = ws.cell(sales_row, col).value if sales_row else None
        cogs_val = ws.cell(cogs_row, col).value if cogs_row else None
        cogs_pct_val = ws.cell(cogs_row, col+1).value if cogs_row else None
        payroll_val = ws.cell(payroll_row, col).value if payroll_row else None
        payroll_pct_val = ws.cell(payroll_row, col+1).value if payroll_row else None
        crew_val = ws.cell(crew_wages_row, col).value if crew_wages_row else None
        crew_pct_val = ws.cell(crew_wages_row, col+1).value if crew_wages_row else None

        budget[store_num][str(period)] = {
            "sales": safe_round(sales_val),
            "cogs": safe_round(cogs_val),
            "cogs_pct": safe_round(cogs_pct_val * 100, 1) if isinstance(cogs_pct_val, (int, float)) else 0,
            "payroll": safe_round(payroll_val),
            "payroll_pct": safe_round(payroll_pct_val * 100, 1) if isinstance(payroll_pct_val, (int, float)) else 0,
            "crew_wages": safe_round(crew_val),
            "crew_wages_pct": safe_round(crew_pct_val * 100, 1) if isinstance(crew_pct_val, (int, float)) else 0,
        }

    p2 = budget[store_num].get("2", {})
    print(f"  {store_num} ({sheet_name:>15}): P2 Sales=${p2.get('sales',0):>12,.2f}  COGS%={p2.get('cogs_pct',0):>5.1f}%  Crew%={p2.get('crew_wages_pct',0):>5.1f}%  TotalPayroll%={p2.get('payroll_pct',0):>5.1f}%")

# Consolidated - Restaurant Operations sheet
# Row 9=Sales, Row 18=COGS, Row 24=Salaries/Wages Crew, Row 28=Total Payroll
ws = wb["Restaurant Operations"]
budget["ALL"] = {"name": "All Stores"}
for period, col in PERIOD_COLS.items():
    budget["ALL"][str(period)] = {
        "sales": safe_round(ws.cell(9, col).value),
        "cogs": safe_round(ws.cell(18, col).value),
        "cogs_pct": safe_round(ws.cell(18, col+1).value * 100, 1) if isinstance(ws.cell(18, col+1).value, (int, float)) else 0,
        "payroll": safe_round(ws.cell(28, col).value),
        "payroll_pct": safe_round(ws.cell(28, col+1).value * 100, 1) if isinstance(ws.cell(28, col+1).value, (int, float)) else 0,
        "crew_wages": safe_round(ws.cell(24, col).value),
        "crew_wages_pct": safe_round(ws.cell(24, col+1).value * 100, 1) if isinstance(ws.cell(24, col+1).value, (int, float)) else 0,
    }

p2 = budget["ALL"]["2"]
print(f"\n  {'ALL':>20}: P2 Sales=${p2['sales']:>12,.2f}  COGS%={p2['cogs_pct']:>5.1f}%  Crew%={p2['crew_wages_pct']:>5.1f}%  TotalPayroll%={p2['payroll_pct']:>5.1f}%")

with open("C:/Users/ascha/OneDrive/Desktop/forage-data/budget_2026.json", "w") as f:
    json.dump(budget, f, indent=2)
print("\nSaved budget_2026.json")
