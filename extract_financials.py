import openpyxl
import json
import os

# Configuration
BASE_DIR = "C:/Users/ascha/OneDrive/Desktop/forage-data"
TMP_DIR = os.path.join(BASE_DIR, "tmp")
OUTPUT_FILE = os.path.join(BASE_DIR, "extracted_data.json")

FILES = {
    "2024": os.path.join(TMP_DIR, "Kitchen Trailing P&L no OH P12.24.xlsx"),
    "2025": os.path.join(TMP_DIR, "Kitchen Trailing P&L no OH P12.25.xlsx"),
}

RESTAURANT_SHEETS = {
    "8001-F": "8001",
    "8002-F": "8002",
    "8003-F": "8003",
    "8004-F": "8004",
    "8005-F": "8005",
    "8006-F": "8006",
    "8007-F": "8007",
    "8008-F": "8008",
    "8009-F": "8009",
    "Restaurant Opera-F": "all_restaurants",
}

METRICS = [
    ("Net Sales", "net_sales", "dollar"),
    ("Total Cost of Goods Sold", "cogs_pct", "pct"),
    ("Total Payroll Expenses", "labor_pct", "pct"),
    ("Total Occupancy", "occupancy_pct", "pct"),
    ("EBITDA", "ebitda_pct", "pct"),
    ("EBITDA", "ebitda_dollars", "dollar"),
]


def period_columns(period_num):
    offset = 12 - period_num
    actual_col = 2 + offset * 2
    pct_col = actual_col + 1
    return actual_col, pct_col


def find_row(ws, search_text, max_row=None):
    if max_row is None:
        max_row = ws.max_row
    for row in range(1, max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            cleaned = str(val).strip()
            if search_text in cleaned:
                return row
    return None


def convert_pct(value):
    if value is None:
        return None
    try:
        v = float(value)
    except (ValueError, TypeError):
        return None
    return round(v * 100, 2)


def convert_dollar(value):
    if value is None:
        return None
    try:
        return round(float(value), 2)
    except (ValueError, TypeError):
        return None


def extract_sheet_data(ws):
    data = {}
    row_cache = {}
    for search_text, output_key, value_type in METRICS:
        cache_key = search_text
        if cache_key not in row_cache:
            row_cache[cache_key] = find_row(ws, search_text)

    for search_text, output_key, value_type in METRICS:
        row = row_cache[search_text]
        if row is None:
            data[output_key] = {f"P{p}": None for p in range(1, 13)}
            continue

        period_data = {}
        for period_num in range(1, 13):
            actual_col, pct_col = period_columns(period_num)
            if value_type == "dollar":
                raw = ws.cell(row=row, column=actual_col).value
                period_data[f"P{period_num}"] = convert_dollar(raw)
            elif value_type == "pct":
                raw = ws.cell(row=row, column=pct_col).value
                period_data[f"P{period_num}"] = convert_pct(raw)
            else:
                period_data[f"P{period_num}"] = None

        data[output_key] = period_data
    return data


def main():
    result = {}

    for year_label, filepath in FILES.items():
        print(f"
" + "=" * 60)
        print(f"Processing {year_label}: {os.path.basename(filepath)}")
        print("=" * 60)

        wb = openpyxl.load_workbook(filepath, data_only=True)
        available_sheets = wb.sheetnames
        print(f"Available sheets: {available_sheets}")

        year_data = {}

        for sheet_name, output_key in RESTAURANT_SHEETS.items():
            if sheet_name not in available_sheets:
                print(f"  SKIP: {sheet_name} not in workbook")
                continue

            ws = wb[sheet_name]
            net_sales_row = find_row(ws, "Net Sales")
            if net_sales_row is None:
                print(f"  SKIP: {sheet_name} -> no Net Sales row found")
                continue

            has_data = False
            for p in range(1, 13):
                ac, _ = period_columns(p)
                v = ws.cell(row=net_sales_row, column=ac).value
                if v is not None and v \!= 0:
                    has_data = True
                    break

            if not has_data:
                print(f"  SKIP: {sheet_name} -> Net Sales all None/0")
                continue

            sheet_data = extract_sheet_data(ws)
            year_data[output_key] = sheet_data

            p12_sales = sheet_data["net_sales"].get("P12")
            p12_cogs = sheet_data["cogs_pct"].get("P12")
            p12_labor = sheet_data["labor_pct"].get("P12")
            p12_ebitda = sheet_data["ebitda_pct"].get("P12")
            print(f"  {sheet_name} -> {output_key}: P12 Sales=${p12_sales}, COGS%={p12_cogs}, Labor%={p12_labor}, EBITDA%={p12_ebitda}")

        result[year_label] = year_data
        wb.close()

    with open(OUTPUT_FILE, "w") as f:
        json.dump(result, f, indent=2)
    print(f"
Saved to: {OUTPUT_FILE}")

    # Detailed summary
    print("

===== DETAILED SUMMARY =====")
    for year_label, year_data in result.items():
        print(f"
--- {year_label} ---")
        print(f"  Stores: {list(year_data.keys())}")
        for store_key, store_data in year_data.items():
            print(f"
  {store_key}:")
            for metric_key, metric_data in store_data.items():
                non_null = sum(1 for v in metric_data.values() if v is not None)
                sample_vals = [f"{k}={v}" for k, v in list(metric_data.items())[:3]]
                print(f"    {metric_key}: {non_null}/12 populated | {', '.join(sample_vals)}")

    # Sample table
    print("

===== SAMPLE DATA TABLE =====")
    for year_label in ["2024", "2025"]:
        print(f"
{year_label}:")
        hdr = f'{"Store":<18} {"Period":<8} {"Net Sales":>12} {"COGS%":>8} {"Labor%":>8} {"Occup%":>8} {"EBITDA%":>8} {"EBITDA$":>12}'
        print(hdr)
        print("-" * 90)
        year_data = result.get(year_label, {})
        for store_key in sorted(year_data.keys()):
            store_data = year_data[store_key]
            for period in ["P1", "P6", "P12"]:
                ns = store_data["net_sales"].get(period)
                cg = store_data["cogs_pct"].get(period)
                lb = store_data["labor_pct"].get(period)
                oc = store_data["occupancy_pct"].get(period)
                eb = store_data["ebitda_pct"].get(period)
                ed = store_data["ebitda_dollars"].get(period)
                ns_s = f"${ns:,.2f}" if ns is not None else "N/A"
                cg_s = f"{cg:.1f}%" if cg is not None else "N/A"
                lb_s = f"{lb:.1f}%" if lb is not None else "N/A"
                oc_s = f"{oc:.1f}%" if oc is not None else "N/A"
                eb_s = f"{eb:.1f}%" if eb is not None else "N/A"
                ed_s = f"${ed:,.2f}" if ed is not None else "N/A"
                print(f"{store_key:<18} {period:<8} {ns_s:>12} {cg_s:>8} {lb_s:>8} {oc_s:>8} {eb_s:>8} {ed_s:>12}")


if __name__ == "__main__":
    main()
